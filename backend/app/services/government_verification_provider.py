import io
import os
import re
import logging
from abc import ABC, abstractmethod
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
import openpyxl

from app.integrations.google_drive import download_file
from app.integrations.firebase.firestore import FirestoreRepository
from app.integrations.firebase.admin import FirebaseAdminManager
from firebase_admin import auth as firebase_auth

logger = logging.getLogger("GovernmentVerification")

def mask_identifier(val: Optional[str]) -> str:
    if not val:
        return "[EMPTY]"
    s = str(val).strip()
    if len(s) <= 4:
        return "****"
    return s[:2] + "*" * (len(s) - 4) + s[-2:]

def normalize_company_name(name: Optional[str]) -> str:
    if not name:
        return ""
    s = str(name).lower().strip()
    s = re.sub(r'^(m/s\.?|m/s|messrs\.?)\s*', '', s)
    s = s.replace("private limited", "pvt ltd").replace("pvt. ltd.", "pvt ltd").replace("pvt ltd.", "pvt ltd")
    s = s.replace("limited", "ltd").replace("ltd.", "ltd")
    s = s.replace("corporation", "corp").replace("corp.", "corp")
    s = s.replace("solutions", "sol").replace("systems", "sys")
    s = re.sub(r'[^\w\s]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _format_date(val: Any) -> str:
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).split(" ")[0].strip()

def _clean_pin(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val)).strip()
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def _is_date_valid(valid_till_str: Any) -> bool:
    if not valid_till_str:
        return False
    if isinstance(valid_till_str, datetime):
        return valid_till_str.year >= 2026

    s = str(valid_till_str).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.year >= 2026
        except ValueError:
            pass
    match = re.search(r'20\d{2}', s)
    if match:
        return int(match.group(0)) >= 2026
    return True


class GovernmentVerificationProvider(ABC):
    """
    Abstract interface for government credential verification.
    Easily swappable with real DigiLocker API in the future without changing consumer code or UI flows.
    """

    @abstractmethod
    def verify_pan(self, pan_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        pass

    @abstractmethod
    def verify_udyam(self, udyam_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        pass

    @abstractmethod
    def verify_gst(self, gst_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        pass

    @abstractmethod
    def verify_oem(self, oem_authorization_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        pass

    @abstractmethod
    def finalize_verification(
        self,
        pan_details: Dict[str, Any],
        udyam_details: Dict[str, Any],
        gst_details: Dict[str, Any],
        oem_details: Dict[str, Any]
    ) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        pass


class GoogleSheetsGovernmentProvider(GovernmentVerificationProvider):
    """
    Simulates DigiLocker / Government API verification using exact Google Sheet datasets.
    """

    PAN_SHEET_ID = os.getenv("GOOGLE_SHEET_PAN_ID", "1S4f3wEKykvZ7LSM8vA129YqNKK9q3dlj")
    UDYAM_SHEET_ID = os.getenv("GOOGLE_SHEET_UDYAM_ID", "1BaLtUC4K7R89UmFK3vHkyGAB-c0HmKNV")
    OEM_SHEET_ID = os.getenv("GOOGLE_SHEET_OEM_ID", "1Vo78YziNt6o4LN4d2BzuwVaPvugFaSOF")
    GST_SHEET_ID = os.getenv("GOOGLE_SHEET_GST_ID", "1EsWtSXWNNNwXcgG1wIhmGIQTJ6Tq89B8")

    _cache: Dict[str, Dict[str, Any]] = {"pan": {}, "udyam": {}, "gst": {}, "oem": {}}
    _last_synced: Optional[datetime] = None

    @classmethod
    def sync_sheets(cls, force: bool = False):
        if cls._last_synced and not force:
            return

        cls._load_sheet("pan", cls.PAN_SHEET_ID, "PAN Details")
        cls._load_sheet("udyam", cls.UDYAM_SHEET_ID, "S.No.,Udyam Registration Number")
        cls._load_sheet("gst", cls.GST_SHEET_ID, "GST_Details")
        cls._load_sheet("oem", cls.OEM_SHEET_ID, "S.No.,Authorization Number,Comp")

        cls._last_synced = datetime.now(timezone.utc)

    @classmethod
    def _load_sheet(cls, category: str, sheet_id: str, target_sheet_name: str):
        try:
            fbytes = download_file(sheet_id)
            wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True)
            
            if target_sheet_name in wb.sheetnames:
                sheet = wb[target_sheet_name]
            else:
                sheet = wb.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return

            headers = [str(c).lower().strip() if c is not None else "" for c in rows[0]]

            for row in rows[1:]:
                if not row or not any(row):
                    continue
                row_dict = {h: v for h, v in zip(headers, row) if h and v is not None}

                if category == "pan":
                    pan_num = cls._get_val(row_dict, ["pan number", "pan", "pannumber"])
                    if pan_num:
                        key = str(pan_num).strip().upper()
                        cls._cache["pan"][key] = {
                            "pan_number": key,
                            "company_name": str(cls._get_val(row_dict, ["company name", "name", "pan holder name"]) or "").strip(),
                            "pan_status": str(cls._get_val(row_dict, ["pan status", "status"]) or "Active").strip(),
                            "date_of_birth": _format_date(cls._get_val(row_dict, ["date of birth", "dob", "date of birth/incorporation"])),
                            "pan_holder_type": str(cls._get_val(row_dict, ["pan holder type", "holder type", "type"]) or "Company").strip(),
                            "aadhaar_number": str(cls._get_val(row_dict, ["aadhaar number", "aadhaar"]) or "").strip(),
                            "aadhaar_linking_status": str(cls._get_val(row_dict, ["pan-aadhaar linking status", "linking status"]) or "Linked").strip(),
                            "remarks": str(cls._get_val(row_dict, ["remarks", "note"]) or "Sample data").strip(),
                            "_pin": _clean_pin(cls._get_val(row_dict, ["pin", "verification pin"]))
                        }

                elif category == "udyam":
                    udyam_num = cls._get_val(row_dict, ["udyam registration number", "udyam number", "registration number"])
                    if udyam_num:
                        key = str(udyam_num).strip().upper()
                        inv_val = cls._get_val(row_dict, ["investment (lakhs)", "investment", "investment (in lakhs)"])
                        turn_val = cls._get_val(row_dict, ["turnover ( lakhs)", "turnover", "turnover (lakhs)"])
                        
                        cls._cache["udyam"][key] = {
                            "udyam_number": key,
                            "company_name": str(cls._get_val(row_dict, ["company name", "enterprise name", "name"]) or "").strip(),
                            "registration_status": str(cls._get_val(row_dict, ["registration status", "status"]) or "Active").strip(),
                            "date_of_registration": _format_date(cls._get_val(row_dict, ["date of registration", "date"])),
                            "enterprise_type": str(cls._get_val(row_dict, ["enterprise type", "type"]) or "Manufacturing").strip(),
                            "investment": f"₹{inv_val} Lakhs" if inv_val is not None and not str(inv_val).startswith("₹") else str(inv_val or "₹150 Lakhs"),
                            "turnover": f"₹{turn_val} Lakhs" if turn_val is not None and not str(turn_val).startswith("₹") else str(turn_val or "₹1200 Lakhs"),
                            "remarks": str(cls._get_val(row_dict, ["remarks", "note"]) or "Sample data").strip(),
                            "_pin": _clean_pin(cls._get_val(row_dict, ["pin", "verification pin"]))
                        }

                elif category == "gst":
                    gstin = cls._get_val(row_dict, ["gstin", "gst number", "gst_number"])
                    if gstin:
                        key = str(gstin).strip().upper()
                        cls._cache["gst"][key] = {
                            "gstin": key,
                            "company_name": str(cls._get_val(row_dict, ["company name", "legal name", "trade name", "name"]) or "").strip(),
                            "registration_status": str(cls._get_val(row_dict, ["registration status", "status"]) or "Active").strip(),
                            "date_of_registration": _format_date(cls._get_val(row_dict, ["date of registration", "date"])),
                            "business_type": str(cls._get_val(row_dict, ["business type", "type"]) or "Manufacturer").strip(),
                            "state": str(cls._get_val(row_dict, ["state", "jurisdiction"]) or "Maharashtra").strip(),
                            "filing_status": str(cls._get_val(row_dict, ["filing status", "compliance status"]) or "Compliant").strip(),
                            "remarks": str(cls._get_val(row_dict, ["remarks", "note"]) or "Sample data").strip(),
                            "_pin": _clean_pin(cls._get_val(row_dict, ["pin", "verification pin"]))
                        }

                elif category == "oem":
                    auth_num = cls._get_val(row_dict, ["authorization number", "auth number", "maf number"])
                    if auth_num:
                        key = str(auth_num).strip().upper()
                        cls._cache["oem"][key] = {
                            "authorization_number": key,
                            "company_name": str(cls._get_val(row_dict, ["company name", "partner name", "name"]) or "").strip(),
                            "authorization_status": str(cls._get_val(row_dict, ["authorization status", "status"]) or "Active").strip(),
                            "date_of_issue": _format_date(cls._get_val(row_dict, ["date of issue", "issue date"])),
                            "valid_till": _format_date(cls._get_val(row_dict, ["valid till", "expiry date"])),
                            "oem_name": str(cls._get_val(row_dict, ["oem name", "oem", "brand"]) or "Dell").strip(),
                            "product_category": str(cls._get_val(row_dict, ["product category", "category"]) or "Hardware").strip(),
                            "remarks": str(cls._get_val(row_dict, ["remarks", "note"]) or "Sample data").strip(),
                            "_pin": _clean_pin(cls._get_val(row_dict, ["pin", "verification pin"]))
                        }
        except Exception as e:
            logger.error(f"Error reading Google Sheet {category} ({sheet_id}): {e}")

    @staticmethod
    def _get_val(d: Dict[str, Any], candidates: List[str]) -> Any:
        for c in candidates:
            for k, v in d.items():
                if c in k:
                    return v
        return None

    def verify_pan(self, pan_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        self.sync_sheets()
        p_clean = pan_number.strip().upper()
        pin_clean = str(pin).strip()

        record = self._cache["pan"].get(p_clean)
        if not record:
            self._log_audit("PAN", p_clean, False, "PAN_NOT_FOUND", "PAN details not found")
            return False, "PAN details could not be verified.", None

        # Compare user PIN with the actual PIN belonging to that specific row
        row_pin = str(record.get("_pin", "")).strip()
        if row_pin != pin_clean:
            self._log_audit("PAN", p_clean, False, "WRONG_PIN", "Incorrect PIN provided")
            return False, "Incorrect PAN verification PIN.", None

        if str(record.get("pan_status", "")).strip().lower() != "active":
            self._log_audit("PAN", p_clean, False, "INACTIVE_PAN", "PAN status is Inactive")
            return False, "PAN details could not be verified.", None

        details = {k: v for k, v in record.items() if not k.startswith("_")}
        self._log_audit("PAN", p_clean, True, "PAN_VERIFIED", "PAN verified successfully")
        return True, "PAN verified successfully", details

    def verify_udyam(self, udyam_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        self.sync_sheets()
        u_clean = udyam_number.strip().upper()
        pin_clean = str(pin).strip()

        record = self._cache["udyam"].get(u_clean)
        if not record:
            self._log_audit("UDYAM", u_clean, False, "UDYAM_NOT_FOUND", "Udyam number not found")
            return False, "Udyam details could not be verified.", None

        # Compare user PIN with the actual PIN belonging to that specific row
        row_pin = str(record.get("_pin", "")).strip()
        if row_pin != pin_clean:
            self._log_audit("UDYAM", u_clean, False, "WRONG_PIN", "Incorrect PIN provided")
            return False, "Incorrect Udyam verification PIN.", None

        if str(record.get("registration_status", "")).strip().lower() != "active":
            self._log_audit("UDYAM", u_clean, False, "INACTIVE_UDYAM", "Udyam status is not Active")
            return False, "Udyam details could not be verified.", None

        details = {k: v for k, v in record.items() if not k.startswith("_")}
        self._log_audit("UDYAM", u_clean, True, "UDYAM_VERIFIED", "Udyam verified successfully")
        return True, "Udyam verified successfully", details

    def verify_gst(self, gst_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        self.sync_sheets()
        g_clean = gst_number.strip().upper()
        pin_clean = str(pin).strip()

        record = self._cache["gst"].get(g_clean)
        if not record:
            self._log_audit("GST", g_clean, False, "GST_NOT_FOUND", "GSTIN not found")
            return False, "GST details could not be verified.", None

        # Compare user PIN with the actual PIN belonging to that specific row
        row_pin = str(record.get("_pin", "")).strip()
        if row_pin != pin_clean:
            self._log_audit("GST", g_clean, False, "WRONG_PIN", "Incorrect PIN provided")
            return False, "Incorrect GST verification PIN.", None

        if str(record.get("registration_status", "")).strip().lower() != "active":
            self._log_audit("GST", g_clean, False, "INACTIVE_GST", "GST status is not Active")
            return False, "GST details could not be verified.", None

        details = {k: v for k, v in record.items() if not k.startswith("_")}
        self._log_audit("GST", g_clean, True, "GST_VERIFIED", "GST verified successfully")
        return True, "GST verified successfully", details

    def verify_oem(self, oem_authorization_number: str, pin: str) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        self.sync_sheets()
        o_clean = oem_authorization_number.strip().upper()
        pin_clean = str(pin).strip()

        record = self._cache["oem"].get(o_clean)
        if not record:
            self._log_audit("OEM", o_clean, False, "OEM_NOT_FOUND", "OEM Authorization number not found")
            return False, "OEM authorization could not be verified.", None

        # Compare user PIN with the actual PIN belonging to that specific row
        row_pin = str(record.get("_pin", "")).strip()
        if row_pin != pin_clean:
            self._log_audit("OEM", o_clean, False, "WRONG_PIN", "Incorrect PIN provided")
            return False, "Incorrect OEM verification PIN.", None

        if not _is_date_valid(record.get("valid_till")) or str(record.get("authorization_status", "")).strip().lower() == "expired":
            self._log_audit("OEM", o_clean, False, "EXPIRED_OEM", "OEM Authorization has expired")
            return False, "OEM authorization has expired.", None

        if str(record.get("authorization_status", "")).strip().lower() != "active":
            self._log_audit("OEM", o_clean, False, "INACTIVE_OEM", "OEM Authorization is not Active")
            return False, "OEM authorization could not be verified.", None

        details = {k: v for k, v in record.items() if not k.startswith("_")}
        self._log_audit("OEM", o_clean, True, "OEM_VERIFIED", "OEM verified successfully")
        return True, "OEM authorization verified successfully", details

    def finalize_verification(
        self,
        pan_details: Dict[str, Any],
        udyam_details: Dict[str, Any],
        gst_details: Dict[str, Any],
        oem_details: Dict[str, Any]
    ) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        c_pan = normalize_company_name(pan_details.get("company_name"))
        c_udyam = normalize_company_name(udyam_details.get("company_name"))
        c_gst = normalize_company_name(gst_details.get("company_name"))
        c_oem = normalize_company_name(oem_details.get("company_name"))

        if not (c_pan and c_udyam and c_gst and c_oem and c_pan == c_udyam == c_gst == c_oem):
            self._log_audit("ALL", pan_details.get("pan_number", ""), False, "COMPANY_MISMATCH", "Documents belong to different companies")
            return False, "The submitted government details do not belong to the same company.", None

        pan_num = pan_details.get("pan_number", "").strip().upper()
        company_id = f"COMP-{pan_num}"
        user_id = f"USR-{pan_num}"
        company_name = pan_details.get("company_name") or gst_details.get("company_name")

        # Sanitize payloads - strictly ensure no PIN is included
        sanitized_pan = {k: v for k, v in pan_details.items() if not k.startswith("_") and k != "pin"}
        sanitized_udyam = {k: v for k, v in udyam_details.items() if not k.startswith("_") and k != "pin"}
        sanitized_gst = {k: v for k, v in gst_details.items() if not k.startswith("_") and k != "pin"}
        sanitized_oem = {k: v for k, v in oem_details.items() if not k.startswith("_") and k != "pin"}

        company_profile = {
            "company_id": company_id,
            "company_name": company_name,
            "pan": sanitized_pan,
            "udyam": sanitized_udyam,
            "gst": sanitized_gst,
            "oem": sanitized_oem,
            "government_verification_status": "VERIFIED",
            "verified_at": datetime.now(timezone.utc).isoformat(),
            "verification_source": "DEMO_GOOGLE_SHEETS",
            "status": "VERIFIED"
        }

        # Store in Firestore (Never storing raw PINs)
        FirestoreRepository.set_document("companies", company_id, company_profile)

        user_profile = {
            "user_id": user_id,
            "company_id": company_id,
            "email": f"bidder.{pan_num.lower()}@gem-bidder.gov.in",
            "role": "BIDDER",
            "government_verified": True,
            "verification_status": "VERIFIED",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        FirestoreRepository.set_document("users", user_id, user_profile)

        # Generate Firebase Custom Auth Token if Firebase Admin is available
        custom_token = None
        try:
            admin_app = FirebaseAdminManager.get_app()
            if admin_app:
                custom_token = firebase_auth.create_custom_token(user_id, {"role": "BIDDER", "company_id": company_id}, app=admin_app).decode("utf-8")
        except Exception as e:
            logger.warning(f"Could not generate Firebase custom token: {e}")

        self._log_audit("ALL", pan_num, True, "FINALIZED", "All government documents verified and company created")

        return True, "Government Details Verified", {
            "verified": True,
            "user_id": user_id,
            "custom_token": custom_token,
            "company": company_profile
        }

    def _log_audit(self, doc_type: str, doc_id: str, success: bool, code: str, reason: str):
        now_ts = datetime.now(timezone.utc)
        event_id = f"AUD-GOV-{now_ts.strftime('%Y%m%d%H%M%S%f')}"
        audit_doc = {
            "event_id": event_id,
            "timestamp": now_ts.isoformat(),
            "event_type": f"GOVERNMENT_{doc_type}_VERIFICATION",
            "result": "SUCCESS" if success else "FAILURE",
            "reason_code": code,
            "reason_description": reason,
            "document_type": doc_type,
            "masked_document_id": mask_identifier(doc_id),
            "verification_source": "DEMO_GOOGLE_SHEETS"
        }
        try:
            FirestoreRepository.set_document("audit_logs", event_id, audit_doc)
        except Exception:
            pass


# Global singleton provider instance
_provider_instance: GovernmentVerificationProvider = GoogleSheetsGovernmentProvider()

def get_government_verification_provider() -> GovernmentVerificationProvider:
    return _provider_instance
