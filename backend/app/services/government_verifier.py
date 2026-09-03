import io
import re
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import openpyxl

from app.integrations.google_drive import download_file, list_drive_folder_pdf_files, get_drive_service
from app.integrations.firebase.firestore import FirestoreRepository
from app.integrations.firebase.admin import FirebaseAdminManager
from firebase_admin import auth as firebase_auth

logger = logging.getLogger("GovernmentVerifier")

def mask_identifier(val: Optional[str]) -> str:
    if not val:
        return "[EMPTY]"
    s = str(val).strip()
    if len(s) <= 4:
        return "****"
    return s[:2] + "*" * (len(s) - 4) + s[-2:]

def normalize_company_name(name: str) -> str:
    if not name:
        return ""
    s = str(name).lower().strip()
    s = re.sub(r'^(m/s\.?|m/s|messrs\.?)\s*', '', s)
    s = s.replace("private limited", "pvt ltd").replace("pvt. ltd.", "pvt ltd").replace("pvt ltd.", "pvt ltd")
    s = s.replace("limited", "ltd").replace("ltd.", "ltd")
    s = s.replace("corporation", "corp").replace("corp.", "corp")
    s = s.replace("solutions", "sol").replace("systems", "sys")
    s = re.sub(r'[^\w\s]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

class GovernmentVerifierService:
    _cached_data: Optional[Dict[str, Any]] = None
    _last_loaded: Optional[datetime] = None
    GOV_FOLDER_ID = "1HUiXTsz1pHPkZl27IZOzeMJEjKrLsss4"

    @classmethod
    def load_government_datasets(cls, force_reload: bool = False) -> Dict[str, Any]:
        if cls._cached_data and not force_reload:
            return cls._cached_data

        datasets = {
            "pan": {},
            "gst": {},
            "udyam": {},
            "oem": {},
            "pin": {},
            "pin_configured": False
        }

        try:
            service = get_drive_service()
            results = service.files().list(
                pageSize=50,
                fields="files(id, name, mimeType)",
                q=f"'{cls.GOV_FOLDER_ID}' in parents and trashed = false"
            ).execute()
            files = results.get("files", [])
            logger.info(f"Scanning {len(files)} files in Google Drive GOVERNMENT DETAILS folder")

            for f in files:
                fname = f.get("name", "").lower()
                fid = f.get("id")
                if fname.endswith(".xlsx") or fname.endswith(".xls") or "spreadsheet" in f.get("mimeType", ""):
                    try:
                        file_bytes = download_file(fid)
                        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                        cls._parse_workbook(wb, datasets)
                    except Exception as e:
                        logger.error(f"Error parsing workbook {f.get('name')}: {e}")

        except Exception as e:
            logger.error(f"Error loading government datasets from Google Drive: {e}")

        # If Drive parsing found records, cache and return
        if datasets["pan"]:
            cls._cached_data = datasets
            cls._last_loaded = datetime.utcnow()
            return datasets

        # Fallback to local default records if Drive network is unreachable
        cls._populate_fallback_records(datasets)
        cls._cached_data = datasets
        cls._last_loaded = datetime.utcnow()
        return datasets

    @classmethod
    def _parse_workbook(cls, wb: openpyxl.Workbook, datasets: Dict[str, Any]):
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            s_lower = sheetname.lower()
            if "instructions" in s_lower:
                continue

            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue

            headers = [str(c).lower().strip() if c is not None else "" for c in rows[0]]

            # Match Sheet Type cleanly
            if "pin" in s_lower or ("verification pin" in headers and "udyam" in headers):
                datasets["pin_configured"] = True
                for row in rows[1:]:
                    if not row or not any(row): continue
                    row_dict = cls._row_to_dict(headers, row)
                    pin_val = cls._find_key(row_dict, ["verification pin", "pin", "verification_pin", "auth pin"])
                    if pin_val is not None:
                        pin_str = str(pin_val).strip()
                        c_name = cls._find_key(row_dict, ["company name", "name"])
                        pan_num = cls._find_key(row_dict, ["pan number", "pan"])
                        gstin = cls._find_key(row_dict, ["gstin", "gst number"])
                        udyam = cls._find_key(row_dict, ["udyam number", "udyam"])
                        oem = cls._find_key(row_dict, ["oem authorization number", "oem number", "auth number"])

                        record = {
                            "pin": pin_str,
                            "company_name": c_name,
                            "pan": str(pan_num).strip().upper() if pan_num else None,
                            "gstin": str(gstin).strip().upper() if gstin else None,
                            "udyam": str(udyam).strip().upper() if udyam else None,
                            "oem": str(oem).strip().upper() if oem else None,
                        }
                        if pan_num:
                            datasets["pin"][str(pan_num).strip().upper()] = record
                        if gstin:
                            datasets["pin"][str(gstin).strip().upper()] = record
                        if c_name:
                            datasets["pin"][normalize_company_name(str(c_name))] = record

            elif "oem" in s_lower or "authorization" in s_lower:
                for row in rows[1:]:
                    if not row or not any(row): continue
                    row_dict = cls._row_to_dict(headers, row)
                    auth_num = cls._find_key(row_dict, ["authorization number", "authorization_number", "auth number", "maf number"])
                    if auth_num:
                        datasets["oem"][str(auth_num).strip().upper()] = {
                            "authorization_number": str(auth_num).strip().upper(),
                            "company_name": cls._find_key(row_dict, ["company name", "partner name", "vendor name", "name"]),
                            "authorization_status": cls._find_key(row_dict, ["authorization status", "status"]),
                            "date_of_issue": str(cls._find_key(row_dict, ["date of issue", "issue date"])),
                            "valid_till": str(cls._find_key(row_dict, ["valid till", "expiry date", "validity"])),
                            "oem_name": cls._find_key(row_dict, ["oem name", "oem", "brand"]),
                            "product_category": cls._find_key(row_dict, ["product category", "category", "scope"])
                        }

            elif "udyam" in s_lower:
                for row in rows[1:]:
                    if not row or not any(row): continue
                    row_dict = cls._row_to_dict(headers, row)
                    udyam_num = cls._find_key(row_dict, ["udyam registration number", "udyam number", "udyam_number", "registration number"])
                    if udyam_num:
                        datasets["udyam"][str(udyam_num).strip().upper()] = {
                            "udyam_number": str(udyam_num).strip().upper(),
                            "company_name": cls._find_key(row_dict, ["company name", "enterprise name", "name"]),
                            "registration_status": cls._find_key(row_dict, ["registration status", "status"]),
                            "date_of_registration": str(cls._find_key(row_dict, ["date of registration", "date", "registration date"])),
                            "enterprise_type": cls._find_key(row_dict, ["enterprise type", "type", "category"]) or "Manufacturing",
                            "investment": str(cls._find_key(row_dict, ["investment", "plant and machinery investment"])),
                            "turnover": str(cls._find_key(row_dict, ["turnover", "annual turnover"]))
                        }

            elif "gst" in s_lower:
                for row in rows[1:]:
                    if not row or not any(row): continue
                    row_dict = cls._row_to_dict(headers, row)
                    gstin = cls._find_key(row_dict, ["gstin", "gst number", "gst_number"])
                    if gstin:
                        datasets["gst"][str(gstin).strip().upper()] = {
                            "gstin": str(gstin).strip().upper(),
                            "company_name": cls._find_key(row_dict, ["company name", "legal name", "trade name", "name"]),
                            "registration_status": cls._find_key(row_dict, ["registration status", "status"]),
                            "date_of_registration": str(cls._find_key(row_dict, ["date of registration", "registration date", "date"])),
                            "business_type": cls._find_key(row_dict, ["business type", "constitution of business", "type"]),
                            "state": cls._find_key(row_dict, ["state", "jurisdiction"]),
                            "filing_status": cls._find_key(row_dict, ["filing status", "compliance status"]) or "Compliant"
                        }

            elif "pan" in s_lower:
                for row in rows[1:]:
                    if not row or not any(row): continue
                    row_dict = cls._row_to_dict(headers, row)
                    pan_num = cls._find_key(row_dict, ["pan number", "pan", "pannumber"])
                    if pan_num:
                        datasets["pan"][str(pan_num).strip().upper()] = {
                            "pan_number": str(pan_num).strip().upper(),
                            "company_name": cls._find_key(row_dict, ["company name", "name", "pan holder name", "holder name"]),
                            "pan_status": cls._find_key(row_dict, ["pan status", "status"]),
                            "date_of_birth": str(cls._find_key(row_dict, ["date of birth/incorporation", "date of birth", "dob", "incorporation date"])),
                            "pan_holder_type": cls._find_key(row_dict, ["pan holder type", "holder type", "type"]),
                            "aadhaar_number": mask_identifier(str(cls._find_key(row_dict, ["aadhaar number", "aadhaar"]))),
                            "aadhaar_linking_status": cls._find_key(row_dict, ["pan-aadhaar linking status", "linking status", "aadhaar status"]),
                            "remarks": cls._find_key(row_dict, ["remarks", "note"]) or "Government Verified Enterprise"
                        }

    @classmethod
    def _row_to_dict(cls, headers: List[str], row: Tuple[Any, ...]) -> Dict[str, Any]:
        d = {}
        for h, v in zip(headers, row):
            if h and v is not None:
                d[h] = v
        return d

    @classmethod
    def _find_key(cls, d: Dict[str, Any], candidates: List[str]) -> Any:
        for c in candidates:
            for k, v in d.items():
                if c in k:
                    return v
        return None

    @classmethod
    def _populate_fallback_records(cls, datasets: Dict[str, Any]):
        datasets["pan"]["ABCDE1234F"] = {
            "pan_number": "ABCDE1234F",
            "company_name": "Nexora Technologies Pvt. Ltd.",
            "pan_status": "Active",
            "date_of_birth": "15-01-2015",
            "pan_holder_type": "Company",
            "aadhaar_number": "1111 **** 3333",
            "aadhaar_linking_status": "Linked",
            "remarks": "Government Verified Enterprise"
        }
        datasets["gst"]["22AAAAA1234A1Z5"] = {
            "gstin": "22AAAAA1234A1Z5",
            "company_name": "Nexora Technologies Pvt. Ltd.",
            "registration_status": "Active",
            "date_of_registration": "01-07-2017",
            "business_type": "Private Limited Company",
            "state": "Maharashtra",
            "filing_status": "Compliant"
        }
        datasets["gst"]["22BBBBB2345B1Z6"] = {
            "gstin": "22BBBBB2345B1Z6",
            "company_name": "BluePeak Solutions Pvt. Ltd.",
            "registration_status": "Active",
            "date_of_registration": "15-08-2017",
            "business_type": "Private Limited Company",
            "state": "Karnataka",
            "filing_status": "Compliant"
        }
        datasets["udyam"]["UDYAM-MH-01-0000001"] = {
            "udyam_number": "UDYAM-MH-01-0000001",
            "company_name": "Nexora Technologies Pvt. Ltd.",
            "registration_status": "Active",
            "date_of_registration": "15-08-2020",
            "enterprise_type": "Manufacturing",
            "investment": "₹5.20 Cr",
            "turnover": "₹24.80 Cr"
        }
        datasets["oem"]["OEM-MH-2026-001"] = {
            "authorization_number": "OEM-MH-2026-001",
            "company_name": "Nexora Technologies Pvt. Ltd.",
            "authorization_status": "Active",
            "date_of_issue": "01-01-2024",
            "valid_till": "31-12-2027",
            "oem_name": "Dell International",
            "product_category": "Hardware & IT Equipment"
        }
        datasets["pin_configured"] = True
        datasets["pin"]["ABCDE1234F"] = {
            "pin": "784920",
            "company_name": "Nexora Technologies Pvt. Ltd.",
            "pan": "ABCDE1234F"
        }

    @classmethod
    def verify_government_details(
        cls,
        pan_number: str,
        gst_number: str,
        udyam_number: str,
        oem_authorization_number: str,
        pin: str
    ) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        """
        Executes strict 5-point government detail verification with cross-document matching.
        """
        datasets = cls.load_government_datasets()

        p_clean = pan_number.strip().upper()
        g_clean = gst_number.strip().upper()
        u_clean = udyam_number.strip().upper()
        o_clean = oem_authorization_number.strip().upper()
        pin_clean = str(pin).strip()

        # 1. PAN Check
        pan_record = datasets["pan"].get(p_clean)
        if not pan_record:
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INVALID_PAN", "PAN not found in dataset")
            return False, "PAN details could not be verified.", None

        if str(pan_record.get("pan_status", "")).strip().lower() != "active":
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INACTIVE_PAN", "PAN status is not Active")
            return False, "PAN details could not be verified.", None

        # 2. GST Check
        gst_record = datasets["gst"].get(g_clean)
        if not gst_record:
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INVALID_GST", "GSTIN not found in dataset")
            return False, "GST details could not be verified.", None

        if str(gst_record.get("registration_status", "")).strip().lower() != "active":
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INACTIVE_GST", "GST status is not Active")
            return False, "GST details could not be verified.", None

        # 3. Udyam Check
        udyam_record = datasets["udyam"].get(u_clean)
        if not udyam_record:
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INVALID_UDYAM", "Udyam number not found in dataset")
            return False, "Udyam details could not be verified.", None

        if str(udyam_record.get("registration_status", "")).strip().lower() != "active":
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INACTIVE_UDYAM", "Udyam status is not Active")
            return False, "Udyam details could not be verified.", None

        # 4. OEM Authorization Check
        oem_record = datasets["oem"].get(o_clean)
        if not oem_record:
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INVALID_OEM", "OEM Authorization number not found")
            return False, "OEM authorization could not be verified.", None

        if str(oem_record.get("authorization_status", "")).strip().lower() != "active":
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INACTIVE_OEM", "OEM status is not Active")
            return False, "OEM authorization could not be verified.", None

        # Check OEM validity date
        valid_till_str = str(oem_record.get("valid_till", "")).strip()
        if not cls._is_oem_valid(valid_till_str):
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "EXPIRED_OEM", "OEM Authorization is expired")
            return False, "OEM authorization could not be verified.", None

        # 5. Cross-Document Company Match
        c_pan = normalize_company_name(pan_record.get("company_name", ""))
        c_gst = normalize_company_name(gst_record.get("company_name", ""))
        c_udyam = normalize_company_name(udyam_record.get("company_name", ""))
        c_oem = normalize_company_name(oem_record.get("company_name", ""))

        if not (c_pan and c_gst and c_udyam and c_oem and c_pan == c_gst == c_udyam == c_oem):
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "COMPANY_MISMATCH", f"Mismatch: pan={c_pan}, gst={c_gst}, udyam={c_udyam}, oem={c_oem}")
            return False, "The government documents do not belong to the same company.", None

        # 6. PIN Check
        if not datasets.get("pin_configured") or not datasets["pin"]:
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "PIN_DATASET_MISSING", "PIN dataset is not configured")
            return False, "Government verification PIN dataset is not configured.", None

        pin_entry = datasets["pin"].get(p_clean) or datasets["pin"].get(g_clean) or datasets["pin"].get(c_pan)
        if not pin_entry or str(pin_entry.get("pin", "")).strip() != pin_clean:
            cls._log_audit(p_clean, g_clean, u_clean, o_clean, False, "INVALID_PIN", "PIN does not match registered dataset")
            return False, "Verification PIN is incorrect.", None

        # All 5 Checks Passed
        company_name = pan_record.get("company_name") or gst_record.get("company_name")
        company_id = f"COMP-{p_clean}"
        user_id = f"USR-{p_clean}"

        company_profile = {
            "company_id": company_id,
            "name": company_name,
            "pan": p_clean,
            "gstin": g_clean,
            "udyam_number": u_clean,
            "oem_authorization_number": o_clean,
            "status": "VERIFIED",
            "verified_at": datetime.utcnow().isoformat(),
            "business_type": gst_record.get("business_type", "Private Limited Company"),
            "state": gst_record.get("state", "Maharashtra"),
            "enterprise_type": udyam_record.get("enterprise_type", "Manufacturing")
        }

        gov_details = {
            "pan": pan_record,
            "gst": gst_record,
            "udyam": udyam_record,
            "oem": oem_record
        }

        # Sync/save to Firestore collections
        cls._save_verified_entities(company_id, user_id, company_profile, p_clean)

        # Record Successful Audit
        cls._log_audit(p_clean, g_clean, u_clean, o_clean, True, "VERIFIED_SUCCESS", "All government records verified and company created")

        # Generate Custom Auth Token if Firebase Admin is available
        custom_token = None
        try:
            admin_app = FirebaseAdminManager.get_app()
            if admin_app:
                custom_token = firebase_auth.create_custom_token(user_id, {"role": "BIDDER", "company_id": company_id}, app=admin_app).decode("utf-8")
        except Exception as e:
            logger.warning(f"Could not generate Firebase custom token: {e}")

        payload = {
            "verified": True,
            "message": "Government details verified successfully",
            "user_id": user_id,
            "custom_token": custom_token,
            "company": company_profile,
            "government_details": gov_details
        }
        return True, "Government details verified successfully", payload

    @classmethod
    def _is_oem_valid(cls, valid_till_str: str) -> bool:
        if not valid_till_str:
            return False
        # Try common date formats
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(valid_till_str, fmt)
                return dt >= datetime(2026, 1, 1)  # Active within current operational period
            except ValueError:
                pass
        # Fallback to year check if string contains year
        match = re.search(r'20\d{2}', valid_till_str)
        if match:
            year = int(match.group(0))
            return year >= 2026
        return True

    @classmethod
    def _save_verified_entities(cls, company_id: str, user_id: str, company_profile: Dict[str, Any], pan: str):
        try:
            # 1. Save to companies collection
            FirestoreRepository.set_document("companies", company_id, company_profile)

            # 2. Save to users collection
            user_doc = {
                "user_id": user_id,
                "email": f"bidder.{pan.lower()}@gem-bidder.gov.in",
                "role": "BIDDER",
                "company_id": company_id,
                "created_at": datetime.utcnow().isoformat(),
                "verified": True
            }
            FirestoreRepository.set_document("users", user_id, user_doc)
        except Exception as e:
            logger.error(f"Error persisting verified company/user to Firestore: {e}")

    @classmethod
    def _log_audit(
        cls,
        pan: str,
        gst: str,
        udyam: str,
        oem: str,
        success: bool,
        code: str,
        reason: str
    ):
        event_id = f"AUD-GOV-{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}"
        audit_doc = {
            "event_id": event_id,
            "timestamp": datetime.utcnow().isoformat(),
            "event_type": "GOVERNMENT_DETAIL_VERIFICATION",
            "verification_method": "FOUR_DOCUMENT_CROSS_MATCH_WITH_PIN",
            "result": "SUCCESS" if success else "FAILURE",
            "reason_code": code,
            "reason_description": reason,
            "documents_checked": ["PAN", "GST", "UDYAM", "OEM_AUTH", "PIN"],
            "masked_identifiers": {
                "pan": mask_identifier(pan),
                "gstin": mask_identifier(gst),
                "udyam": mask_identifier(udyam),
                "oem_authorization": mask_identifier(oem)
            }
            # Notice: Never store raw or masked PIN
        }
        try:
            FirestoreRepository.set_document("audit_logs", event_id, audit_doc)
        except Exception as e:
            logger.error(f"Error recording audit log: {e}")
