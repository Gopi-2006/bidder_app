import io
import sys
import openpyxl
from app.integrations.google_drive import download_file

sys.stdout.reconfigure(encoding='utf-8')
oem_sheet_id = "1Vo78YziNt6o4LN4d2BzuwVaPvugFaSOF"
fbytes = download_file(oem_sheet_id)
wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True)
print(f"Sheetnames: {wb.sheetnames}")
sheet = wb.active
rows = list(sheet.iter_rows(values_only=True))
print("Headers:", rows[0])
for row in rows[1:4]:
    print("Row:", row)
