import os
import openpyxl
from datetime import datetime

data_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "app", "data", "government_datasets"))
os.makedirs(data_dir, exist_ok=True)

# 1. PAN.xlsx
wb_pan = openpyxl.Workbook()
ws_pan = wb_pan.active
ws_pan.title = "PAN Details"
ws_pan.append(["S.No.", "PAN Number", "Company Name", "PAN Status", "Date of Birth", "PAN Holder Type", "Aadhaar Number", "PAN-Aadhaar Linking Status", "Remarks", "pin"])
pan_rows = [
    [1, "ABCDE1234F", "Nexora Technologies Pvt. Ltd.", "Active", "1985-01-15", "Company", "1111 2222 3333", "Linked", "Verified Enterprise", "123456"],
    [2, "BCDEF2345G", "BluePeak Solutions Pvt. Ltd.", "Active", "1985-06-01", "Company", "1111 2222 3334", "Linked", "Verified Enterprise", "123456"],
    [3, "CDEFG3456H", "CloudNest Innovations Pvt. Ltd.", "Active", "1985-10-16", "Company", "1111 2222 3335", "Linked", "Verified Enterprise", "123456"],
    [4, "DEFGH4567J", "TechVanta Systems Pvt. Ltd.", "Active", "1986-03-02", "Company", "1111 2222 3336", "Not Linked", "Verified Enterprise", "123456"],
    [5, "EFGHJ5678K", "ByteBridge Technologies Pvt. Ltd", "Active", "1986-07-17", "Company", "1111 2222 3337", "Linked", "Verified Enterprise", "123456"],
    [6, "FGHJK6789L", "NovaCore Solutions Pvt. Ltd.", "Inactive", "1986-12-01", "Company", "1111 2222 3338", "Not Applicable", "Deactivated", "123456"],
]
for r in pan_rows: ws_pan.append(r)
wb_pan.save(os.path.join(data_dir, "PAN.xlsx"))

# 2. UDYAM.xlsx
wb_udyam = openpyxl.Workbook()
ws_udyam = wb_udyam.active
ws_udyam.title = "Udyam Registration"
ws_udyam.append(["S.No.", "Udyam Registration Number", "Company Name", "Registration Status", "Date of Registration", "Enterprise Type", "Investment (lakhs)", "Turnover ( Lakhs)", "Remarks", "pin"])
udyam_rows = [
    [1, "UDYAM-MH-01-0000001", "Nexora Technologies Pvt. Ltd.", "Active", "2020-07-01", "Manufacturing", 150, 1200, "Verified MSME", "123456"],
    [2, "UDYAM-KA-02-0000002", "BluePeak Solutions Pvt. Ltd.", "Active", "2020-08-15", "Service", 80, 950, "Verified MSME", "123456"],
    [3, "UDYAM-TN-03-0000003", "CloudNest Innovations Pvt. Ltd.", "Active", "2020-09-20", "Manufacturing", 200, 1800, "Verified MSME", "123456"],
    [4, "UDYAM-UP-04-0000004", "TechVanta Systems Pvt. Ltd.", "Active", "2020-10-10", "Manufacturing", 120, 800, "Verified MSME", "123456"],
    [5, "UDYAM-GJ-05-0000005", "ByteBridge Technologies Pvt. Ltd", "Active", "2020-11-14", "Service", 60, 500, "Verified MSME", "123456"],
    [6, "UDYAM-RJ-06-0000006", "NovaCore Solutions Pvt. Ltd.", "Inactive", "2020-12-20", "Manufacturing", 90, 650, "Suspended", "123456"],
]
for r in udyam_rows: ws_udyam.append(r)
wb_udyam.save(os.path.join(data_dir, "UDYAM.xlsx"))

# 3. GST.xlsx
wb_gst = openpyxl.Workbook()
ws_gst = wb_gst.active
ws_gst.title = "GST Details"
ws_gst.append(["S.No.", "GSTIN", "Company Name", "Registration Status", "Date of Registration", "Business Type", "State", "Filing Status", "Remarks", "pin"])
gst_rows = [
    [1, "22AAAAA1234A1Z5", "Nexora Technologies Pvt. Ltd.", "Active", "2017-07-01", "Private Limited Company", "Maharashtra", "Compliant", "Verified GSTIN", "123456"],
    [2, "29BBBBB2345B1Z4", "BluePeak Solutions Pvt. Ltd.", "Active", "2017-07-01", "Private Limited Company", "Karnataka", "Compliant", "Verified GSTIN", "123456"],
    [3, "33CCCCC3456C1Z3", "CloudNest Innovations Pvt. Ltd.", "Active", "2017-07-01", "Private Limited Company", "Tamil Nadu", "Compliant", "Verified GSTIN", "123456"],
    [4, "09DDDDD4567D1Z2", "TechVanta Systems Pvt. Ltd.", "Active", "2018-08-15", "Private Limited Company", "Uttar Pradesh", "Compliant", "Verified GSTIN", "123456"],
    [5, "24EEEEE5555E1Z1", "ByteBridge Technologies Pvt. Ltd", "Active", "2019-09-10", "Private Limited Company", "Gujarat", "Compliant", "Verified GSTIN", "123456"],
    [6, "08FFFFF6666F1Z0", "NovaCore Solutions Pvt. Ltd.", "Inactive", "2017-07-01", "Private Limited Company", "Rajasthan", "Non-Compliant", "Suspended", "123456"],
]
for r in gst_rows: ws_gst.append(r)
wb_gst.save(os.path.join(data_dir, "GST.xlsx"))

# 4. OEM.xlsx
wb_oem = openpyxl.Workbook()
ws_oem = wb_oem.active
ws_oem.title = "OEM Authorization"
ws_oem.append(["S.No.", "Authorization Number", "Company Name", "Authorization Status", "Date of Issue", "Valid Till", "OEM Name", "Product Category", "Remarks", "pin"])
oem_rows = [
    [1, "OEM-MH-2026-001", "Nexora Technologies Pvt. Ltd.", "Active", "2026-01-01", "2027-12-31", "Dell", "Hardware", "Verified Authorization", "123456"],
    [2, "OEM-KA-2026-002", "BluePeak Solutions Pvt. Ltd.", "Active", "2026-02-01", "2027-01-31", "Cisco", "Networking", "Verified Authorization", "123456"],
    [3, "OEM-TN-2026-003", "CloudNest Innovations Pvt. Ltd.", "Active", "2026-03-01", "2027-02-28", "Microsoft", "Software", "Verified Authorization", "123456"],
    [4, "OEM-UP-2026-004", "TechVanta Systems Pvt. Ltd.", "Active", "2026-04-01", "2027-03-31", "HP", "Hardware", "Verified Authorization", "123456"],
    [5, "OEM-GJ-2026-005", "ByteBridge Technologies Pvt. Ltd", "Active", "2026-05-01", "2027-04-30", "Siemens", "Industrial", "Verified Authorization", "123456"],
    [6, "OEM-MH-2022-999", "Expired OEM Partner", "Expired", "2021-01-01", "2023-01-01", "Legacy OEM", "Obsolete Hardware", "Expired", "123456"],
]
for r in oem_rows: ws_oem.append(r)
wb_oem.save(os.path.join(data_dir, "OEM.xlsx"))

print(f"Created all 4 local Excel datasets in {data_dir}")
