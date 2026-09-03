import io
import sys
import openpyxl
from app.integrations.google_drive import download_file

sys.stdout.reconfigure(encoding='utf-8')

oem_sheet_id = "1Vo78YziNt6o4LN4d2BzuwVaPvugFaSOF"
fbytes = download_file(oem_sheet_id)
wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True)
ws = wb.active
for r in range(1, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {r}: {row_vals}")
