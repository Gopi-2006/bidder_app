import io
import sys
import openpyxl
from app.integrations.google_drive import download_file, get_drive_service

sys.stdout.reconfigure(encoding='utf-8')

sheet_ids = {
    "PAN": "1S4f3wEKykvZ7LSM8vA129YqNKK9q3dlj",
    "UDYAM": "1BaLtUC4K7R89UmFK3vHkyGAB-c0HmKNV",
    "OEM": "1Vo78YziNt6o4LN4d2BzuwVaPvugFaSOF",
    "GST": "1EsWtSXWNNNwXcgG1wIhmGIQTJ6Tq89B8"
}

for doc_type, sheet_id in sheet_ids.items():
    print(f"\n==================== {doc_type} (ID: {sheet_id}) ====================")
    try:
        content = download_file(sheet_id)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        ws = wb.active
        for r in range(1, min(ws.max_row + 1, 10)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            print(f"Row {r}: {row_vals}")
    except Exception as e:
        print(f"Error inspecting {doc_type}: {e}")
