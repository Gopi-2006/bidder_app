import io
import sys
import openpyxl
from googleapiclient.http import MediaIoBaseUpload
from app.integrations.google_drive import get_drive_service, download_file

sys.stdout.reconfigure(encoding='utf-8')
service = get_drive_service()

pan_fid = "1S4f3wEKykvZ7LSM8vA129YqNKK9q3dlj"
udyam_fid = "1BaLtUC4K7R89UmFK3vHkyGAB-c0HmKNV"
oem_fid = "1Vo78YziNt6o4LN4d2BzuwVaPvugFaSOF"

def inspect_file(fid, label):
    try:
        content = download_file(fid)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        print(f"=== {label} (Sheets: {wb.sheetnames}) ===")
        ws = wb.active
        for r in range(1, min(ws.max_row + 1, 10)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            print(f"Row {r}: {row_vals}")
    except Exception as e:
        print(f"Error inspecting {label}: {e}")

inspect_file(pan_fid, "PAN.Xlsx")
inspect_file(udyam_fid, "UDYAM.Xlsx")
inspect_file(oem_fid, "OEM.xlsx")
